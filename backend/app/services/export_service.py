from datetime import datetime, time
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app.models.attendance import AttendanceRecord
from app.models.user import User
from app.services.attendance_service import month_bounds, parse_month


DAY_TYPE_HE: dict[str, str] = {
    "work": "עבודה",
    "vacation": "חופשה",
    "sick": "מחלה",
    "reserve": "מילואים",
    "holiday": "חג",
    "other_absence": "היעדרות אחרת",
    "full_day_activity": "פעילות יום מלא",
}

STATUS_HE: dict[str, str] = {
    "draft": "טיוטה",
    "submitted": "הוגש",
    "approved": "אושר",
    "rejected": "נדחה",
}


def _hours(check_in: time | None, check_out: time | None) -> float:
    if check_in is None or check_out is None:
        return 0.0
    base = datetime(2000, 1, 1)
    delta = datetime.combine(base, check_out) - datetime.combine(base, check_in)
    return round(delta.total_seconds() / 3600, 2)


def _row(
    user: User,
    rec: AttendanceRecord,
    day_type: str,
    with_hours: bool,
) -> list:
    return [
        user.full_name,
        user.email,
        user.department or "",
        rec.date.isoformat(),
        DAY_TYPE_HE.get(day_type, day_type),
        rec.check_in.isoformat() if (with_hours and rec.check_in) else "",
        rec.check_out.isoformat() if (with_hours and rec.check_out) else "",
        _hours(rec.check_in, rec.check_out) if with_hours else 0.0,
        STATUS_HE.get(rec.status, rec.status),
        rec.note or "",
    ]


def export_month_xlsx(db: Session, month: str) -> tuple[bytes, str]:
    month_first = parse_month(month)
    start, end = month_bounds(month_first)

    rows = (
        db.query(AttendanceRecord, User)
        .join(User, User.id == AttendanceRecord.user_id)
        .filter(AttendanceRecord.date >= start, AttendanceRecord.date <= end)
        .order_by(User.full_name, AttendanceRecord.date)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = month_first.strftime("%Y-%m")
    ws.sheet_view.rightToLeft = True

    headers = [
        "עובד",
        "אימייל",
        "מחלקה",
        "תאריך",
        "סוג",
        "כניסה",
        "יציאה",
        "שעות",
        "סטטוס",
        "הערה",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F4F8A")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for rec, user in rows:
        if rec.partial_secondary_type is None:
            ws.append(_row(user, rec, rec.day_type, True))
        else:
            primary_with_hours = rec.day_type == "work"
            secondary_with_hours = rec.partial_secondary_type == "work"
            # If neither side is work but the record happens to carry hours,
            # attach them to the primary row so they're not silently dropped.
            if not primary_with_hours and not secondary_with_hours:
                primary_with_hours = True
            ws.append(_row(user, rec, rec.day_type, primary_with_hours))
            ws.append(
                _row(user, rec, rec.partial_secondary_type, secondary_with_hours)
            )

    for col_idx, header in enumerate(headers, start=1):
        max_len = max(
            [len(str(header))]
            + [
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(2, ws.max_row + 1)
            ]
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
            max_len + 2, 40
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"cybrella-time-{month_first.strftime('%Y-%m')}.xlsx"
    return buf.read(), filename
